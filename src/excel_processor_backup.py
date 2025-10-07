"""
Processador de arquivos Excel
Responsável por carregar, processar e salvar planilhas Excel
Implementa a lógica do VBA original em Python moderno
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
from typing import Optional, Tuple, List, Dict
from calculator import CalculationEngine

class ExcelProcessor:
    def __init__(self, file_path: str):
        """
        Inicializa o processador Excel
        
        Args:
            file_path: Caminho para o arquivo Excel
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.df = None
        self.original_df = None
        self.calculation_engine = CalculationEngine()
        self.file_format = "desconhecido"  # "novo", "antigo" ou "desconhecido"
        self.lojas_disponiveis = []  # Lista de códigos de lojas encontradas
        self.load_excel_file()
        
    def load_excel_file(self):
        """Carrega o arquivo Excel"""
        try:
            print(f"\n=== INICIANDO ANÁLISE DO ARQUIVO ===")
            print(f"Arquivo: {self.file_path}")
            
            # Carrega o workbook com openpyxl para manter formatação
            self.workbook = load_workbook(self.file_path)
            print(f"Planilhas encontradas: {self.workbook.sheetnames}")
            
            # Pega a segunda planilha (Plan2 do VBA original) ou primeira se só houver uma
            if len(self.workbook.sheetnames) >= 2:
                sheet_name = self.workbook.sheetnames[1]
                print(f"Usando segunda planilha: {sheet_name}")
            else:
                sheet_name = self.workbook.sheetnames[0]
                print(f"Usando primeira planilha: {sheet_name}")
                
            self.worksheet = self.workbook[sheet_name]
            
            # Primeiro carrega tudo sem header para analisar a estrutura
            temp_df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            print(f"Dados brutos carregados: {len(temp_df)} linhas x {len(temp_df.columns)} colunas")
            
            # Debug das primeiras 20 linhas para entender a estrutura
            print(f"\n=== ANÁLISE DAS PRIMEIRAS 20 LINHAS ===")
            for i in range(min(20, len(temp_df))):
                row_data = []
                for j in range(min(10, len(temp_df.columns))):  # Primeiras 10 colunas
                    cell = temp_df.iloc[i, j]
                    if pd.notna(cell):
                        row_data.append(f"Col{j}:'{str(cell)[:30]}'")
                
                if row_data:  # Só mostra linhas que não estão vazias
                    print(f"Linha {i}: {' | '.join(row_data)}")
            
            # Procura pela linha que contém os headers (Tipo, Codigo, Descricao, etc.)
            header_row = self.find_header_row(temp_df)
            print(f"\nHeader encontrado na linha: {header_row}")
            
            if header_row >= 0:
                # Carrega novamente usando a linha correta como header
                self.df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row)
                # Remove linhas vazias
                self.df = self.df.dropna(how='all').reset_index(drop=True)
                print(f"Dados com header carregados: {len(self.df)} linhas x {len(self.df.columns)} colunas")
                print(f"Colunas encontradas: {list(self.df.columns)[:10]}...")  # Primeiras 10 colunas
            else:
                # Se não encontrar header, usa dados brutos
                self.df = temp_df.dropna(how='all').reset_index(drop=True)
                print(f"Usando dados brutos (sem header): {len(self.df)} linhas x {len(self.df.columns)} colunas")
            
            self.original_df = self.df.copy()
            
            # Debug das primeiras linhas dos dados finais
            print(f"\n=== DADOS FINAIS CARREGADOS ===")
            if len(self.df) > 0 and len(self.df.columns) > 0:
                print(f"Primeira coluna: '{self.df.columns[0] if hasattr(self.df, 'columns') else 'Sem nome'}'")
                for i in range(min(5, len(self.df))):
                    primeira_coluna = self.df.iloc[i, 0] if len(self.df.columns) > 0 else "N/A"
                    print(f"  Linha {i}: '{primeira_coluna}'")
            
            # Analisa se é formato novo (com lojas)
            self.analyze_file_format()
            
        except Exception as e:
            print(f"ERRO ao carregar arquivo: {str(e)}")
            raise Exception(f"Erro ao carregar arquivo Excel: {str(e)}")
            
    def analyze_file_format(self):
        """Analisa se o arquivo está no formato novo ou antigo"""
        print(f"\n=== ANÁLISE DO FORMATO DO ARQUIVO ===")
        
        if len(self.df) == 0:
            print("❌ DataFrame vazio - não foi possível analisar formato")
            return
            
        # Procura por indicadores de formato novo (lojas entre parênteses)
        lojas_encontradas = []
        linhas_com_loja = []
        
        for idx, row in self.df.iterrows():
            primeira_coluna = row.iloc[0] if len(row) > 0 else ""
            
            if pd.notna(primeira_coluna):
                cell_str = str(primeira_coluna)
                
                # Procura padrão "Loja: XX-XXX (XXX)"
                if "Loja:" in cell_str and "(" in cell_str and ")" in cell_str:
                    # Extrai o código da loja entre parênteses
                    start = cell_str.find("(")
                    end = cell_str.find(")", start)
                    if start != -1 and end != -1:
                        loja_codigo = cell_str[start+1:end]
                        lojas_encontradas.append(loja_codigo)
                        linhas_com_loja.append(idx)
                        print(f"✓ Loja encontrada na linha {idx}: '{loja_codigo}' (texto completo: '{cell_str}')")
        
        print(f"Total de lojas encontradas: {len(lojas_encontradas)}")
        print(f"Códigos das lojas: {lojas_encontradas}")
        
        if len(lojas_encontradas) > 0:
            print("✓ Formato NOVO detectado (com lojas entre parênteses)")
            self.file_format = "novo"
            self.lojas_disponiveis = lojas_encontradas
        else:
            print("✓ Formato ANTIGO detectado (sem lojas entre parênteses)")
            self.file_format = "antigo"
            self.lojas_disponiveis = []
            
        # Procura por dados de produtos
        produtos_encontrados = 0
        for idx, row in self.df.iterrows():
            # Verifica se a linha parece ser de produto (tem código e descrição)
            if len(row) >= 3:
                col1 = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                col2 = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                col3 = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
                
                # Se não é linha de loja e tem dados nas primeiras colunas
                if not ("Loja:" in col1) and len(col2) > 3 and len(col3) > 5:
                    produtos_encontrados += 1
                    if produtos_encontrados <= 5:  # Mostra os primeiros 5
                        print(f"  Produto encontrado linha {idx}: '{col1}' | '{col2}' | '{col3}'")
                    
        print(f"Total de produtos/itens encontrados: {produtos_encontrados}")
        
        if produtos_encontrados == 0:
            print("❌ PROBLEMA: Nenhum produto encontrado no arquivo!")
            print("   Verifique se os dados estão na planilha correta")
            
    def find_header_row(self, df: pd.DataFrame) -> int:
        """
        Encontra a linha que contém os cabeçalhos das colunas
        No novo formato: Zona, Código do Item, Código Secundário, Item, etc.
        """
        print(f"\n=== PROCURANDO LINHA DE CABEÇALHO ===")
        
        # Palavras-chave para formato antigo e novo
        keywords_antigo = ['tipo', 'codigo', 'descricao', 'entrada', 'vendas', 'saldo']
        keywords_novo = ['zona', 'código do item', 'código secundário', 'item', 'vendas', 'entradas', 'estoque', 'requisições']
        
        for row_idx in range(min(20, len(df))):  # Verifica primeiras 20 linhas
            row_values = []
            
            # Pega valores das primeiras colunas da linha
            for col_idx in range(min(10, len(df.columns))):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value):
                    row_values.append(str(cell_value).lower().strip())
            
            # Verifica se esta linha contém palavras-chave de header
            keyword_count_antigo = sum(1 for keyword in keywords_antigo if any(keyword in val for val in row_values))
            keyword_count_novo = sum(1 for keyword in keywords_novo if any(keyword in val for val in row_values))
            
            print(f"Linha {row_idx}: {row_values[:5]}... - Antigo:{keyword_count_antigo}, Novo:{keyword_count_novo}")
            
            # Se encontrou pelo menos 2 keywords de qualquer formato
            if keyword_count_antigo >= 2 or keyword_count_novo >= 2:
                formato = "antigo" if keyword_count_antigo >= keyword_count_novo else "novo"
                print(f"✓ Header encontrado na linha {row_idx} (formato {formato})")
                return row_idx
                
        print("❌ Nenhum header encontrado")
        return -1  # Não encontrou header
            
    def find_data_start_row(self, df: pd.DataFrame) -> int:
        """
        Encontra onde começam os dados reais no Excel
        Procura por padrões típicos como 'Vendas' ou dados numéricos
        
        Args:
            df: DataFrame completo
            
        Returns:
            Índice da linha onde começam os dados (geralmente linha 8, índice 7)
        """
        for idx, row in df.iterrows():
            # Verifica se a primeira coluna contém 'Vendas' ou similar
            if pd.notna(row.iloc[0]):
                cell_value = str(row.iloc[0]).strip().lower()
                if any(keyword in cell_value for keyword in ['vendas', 'produto', 'item', 'codigo']):
                    print(f"Dados encontrados a partir da linha {idx + 1}")
                    return idx
                    
            # Se chegou na linha 15 e ainda não encontrou, assume que começa na linha 8
            if idx >= 15:
                print("Assumindo início dos dados na linha 8")
                return 7  # Linha 8 (índice 7)
                
        # Se não encontrou nada, retorna linha 8 como padrão
        print("Usando linha 8 como padrão para início dos dados")
        return 7
            
    def get_data_preview(self) -> pd.DataFrame:
        """Retorna preview dos dados para visualização"""
        if self.df is not None and not self.df.empty:
            # Tenta fazer a transposição para o formato antigo
            try:
                print("\n🔄 Tentando transposição para preview...")
                transposed_df = self.transpose_to_old_format()
                
                if not transposed_df.empty:
                    print(f"✅ Preview: Dados transpostos com sucesso - {len(transposed_df)} linhas")
                    return transposed_df
                else:
                    print("⚠️ Transposição resultou em DataFrame vazio")
                    
            except Exception as e:
                print(f"❌ Erro na transposição: {str(e)}")
                
            # Se transposição falhar, mostra dados originais
            print("📊 Mostrando dados originais como fallback")
            preview_df = self.df.copy()
            
            # Adiciona informações de debug
            print(f"Preview original: {len(preview_df)} linhas, {len(preview_df.columns)} colunas")
            if len(preview_df) > 0:
                print(f"Primeira linha: {list(preview_df.iloc[0])}")
                
            return preview_df
        else:
            print("❌ DataFrame vazio ou None")
            return pd.DataFrame()
        
    def find_product_rows(self) -> List[int]:
        """
        Encontra todas as linhas que contêm dados de produtos (não são cabeçalhos de loja)
        No novo formato, identifica linhas com código de produto válido
        
        Returns:
            Lista de índices das linhas que contêm dados de produtos
        """
        product_rows = []
        
        if self.df is None or self.df.empty:
            print("DataFrame vazio - nenhuma linha de produto encontrada")
            return product_rows
        
        print("Procurando por linhas de produtos...")
        
        # Identifica colunas importantes
        cols = list(self.df.columns)
        zona_col = cols[0] if len(cols) > 0 else None  # Coluna A: Zona
        codigo_col = cols[1] if len(cols) > 1 else None  # Coluna B: Código do Item
        vendas_col = cols[5] if len(cols) > 5 else None  # Coluna F: Vendas
        
        print(f"Debug: Colunas identificadas - Zona: {zona_col}, Código: {codigo_col}, Vendas: {vendas_col}")
        
        for idx, row in self.df.iterrows():
            zona_val = row[zona_col] if zona_col else None
            codigo_val = row[codigo_col] if codigo_col else None
            
            # Verifica se é uma linha de produto (tem código e não é cabeçalho de loja)
            if pd.notna(codigo_val) and pd.notna(zona_val):
                zona_str = str(zona_val).strip()
                codigo_str = str(codigo_val).strip()
                
                # Não é cabeçalho de loja se não contém "Loja:" e tem código numérico
                if not zona_str.startswith("Loja:") and codigo_str.isdigit():
                    product_rows.append(idx)
                    print(f"✓ Linha de produto encontrada: {idx} - {zona_str} - {codigo_str}")
                    
        print(f"Total de linhas de produtos encontradas: {len(product_rows)}")
        return product_rows
        
    def find_vendas_rows(self) -> List[int]:
        """
        Compatibilidade com código antigo - agora retorna linhas de produtos
        """
        return self.find_product_rows()
        
    def check_suggestion_exists(self) -> bool:
        """
        Verifica se já existe uma linha de 'Sugestão' nos dados
        
        Returns:
            True se encontrar 'Sugestão', False caso contrário
        """
    def check_suggestion_exists(self) -> bool:
        """
        Verifica se já existe uma coluna de 'Sugestão' com dados
        No novo formato, verifica se a coluna S já tem valores
        
        Returns:
            True se encontrar dados de sugestão, False caso contrário
        """
        if self.df is None or self.df.empty:
            return False
            
        print("Verificando se já existem sugestões...")
        
        # No novo formato, a sugestão fica na coluna S (índice 18)
        if len(self.df.columns) > 18:
            suggestion_col = self.df.columns[18]
            
            # Verifica se há valores não vazios na coluna de sugestão
            for idx, row in self.df.iterrows():
                valor = row[suggestion_col]
                if pd.notna(valor) and str(valor).strip() != "" and str(valor).strip() != "0":
                    print(f"Sugestão encontrada na linha {idx}: '{valor}'")
                    return True
                    
        print("Nenhuma sugestão encontrada - OK para calcular")
        return False
        
    def calculate_suggestions(self, percentage: float, progress_callback=None) -> bool:
        """
        Aplica a porcentagem nos dados transpostos (formato antigo)
        """
        try:
            print(f"\n🧮 INICIANDO CÁLCULO DE SUGESTÕES ({percentage}%)")
            
            # Primeiro faz a transposição
            transposed_df = self.transpose_to_old_format()
            
            if transposed_df.empty:
                raise Exception("Não foi possível transpor os dados para o formato antigo")
                
            print(f"📊 Dados transpostos: {len(transposed_df)} linhas")
            
            # Aplica porcentagem apenas nas linhas de Vendas
            modified_data = []
            vendas_count = 0
            
            for idx, row in transposed_df.iterrows():
                if row.get('Tipo') == 'Vendas':
                    # Cria linha de sugestão
                    suggestion_row = row.copy()
                    suggestion_row['Tipo'] = 'Sugestão'
                    
                    # Aplica porcentagem em todas as colunas de lojas
                    total_sugestao = 0
                    for col in suggestion_row.index:
                        if col not in ['Tipo', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 
                                     'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada', 'Qtde Total']:
                            valor_original = suggestion_row[col]
                            if pd.notna(valor_original) and isinstance(valor_original, (int, float)):
                                novo_valor = valor_original * (1 + percentage / 100)
                                suggestion_row[col] = novo_valor
                                total_sugestao += novo_valor
                                
                    suggestion_row['Qtde Total'] = total_sugestao
                    
                    # Adiciona linha original e sugestão
                    modified_data.append(row.to_dict())
                    modified_data.append(suggestion_row.to_dict())
                    vendas_count += 1
                    
                    if progress_callback:
                        progress = int((vendas_count / (len(transposed_df) // 2)) * 100)
                        progress_callback.emit(progress)
                else:
                    # Adiciona linha de entrada sem modificação
                    modified_data.append(row.to_dict())
                    
            self.df = pd.DataFrame(modified_data)
            
            print(f"✅ Cálculo concluído: {vendas_count} sugestões criadas")
            return True
            
        except Exception as e:
            print(f"❌ Erro durante cálculo: {str(e)}")
            raise Exception(f"Erro durante cálculo: {str(e)}")
            
    def insert_suggestion_row(self, vendas_idx: int, percentage: float):
        """
        Insere linha de sugestão após uma linha de vendas
        
        Args:
            vendas_idx: Índice da linha de vendas
            percentage: Porcentagem de ajuste
        """
        # Cria nova linha
        suggestion_row = self.df.iloc[vendas_idx].copy()
        suggestion_row.iloc[0] = "Sugestão"
        
        # Copia valores das colunas 2-10 (índices 1-9 em Python)
        for col_idx in range(1, 10):
            if col_idx < len(suggestion_row):
                suggestion_row.iloc[col_idx] = self.df.iloc[vendas_idx, col_idx]
                
        # Aplica porcentagem nas colunas 11-28 (índices 10-27 em Python)
        soma = 0
        for col_idx in range(10, 28):
            if col_idx < len(suggestion_row):
                original_value = self.df.iloc[vendas_idx, col_idx]
                if pd.notna(original_value) and isinstance(original_value, (int, float)):
                    new_value = original_value * (1 + percentage / 100)
                    suggestion_row.iloc[col_idx] = new_value
                    soma += new_value
                    
        # Define soma na coluna 29 (índice 28)
        if len(suggestion_row) > 28:
            suggestion_row.iloc[28] = soma
            
        # Insere a linha no DataFrame
        self.df = pd.concat([
            self.df.iloc[:vendas_idx + 1],
            pd.DataFrame([suggestion_row]),
            self.df.iloc[vendas_idx + 1:]
        ]).reset_index(drop=True)
        
    def apply_formatting(self):
        """Aplica formatação às colunas - inclui cabeçalho de Sugestão"""
        if self.df is None or self.df.empty:
            return
            
        # Garante que existe coluna de Sugestão
        if len(self.df.columns) < 19:
            # Adiciona colunas vazias até chegar na posição S
            while len(self.df.columns) < 19:
                new_col_name = f"Col_{len(self.df.columns)}"
                self.df[new_col_name] = ""
                
        # Renomeia a coluna 18 (índice 18) para "Sugestão"
        cols = list(self.df.columns)
        cols[18] = "Sugestão"
        self.df.columns = cols
        
        print("Coluna de Sugestão criada/atualizada")
        
    def transform_to_table(self):
        """Equivalente à função TransformarEmTabela do VBA"""
        # Implementação simplificada - no contexto real seria mais complexa
        # Por enquanto, apenas garantimos que os dados estão organizados
        pass
        
    def sum_values_by_criteria(self):
        """Equivalente à função SomarValoresPorCritério do VBA"""
        # Implementação simplificada
        # Aqui faríamos somas condicionais conforme necessário
        pass
        
    def export_results(self, output_path: str):
        """
        Exporta os resultados para um novo arquivo Excel
        
        Args:
            output_path: Caminho do arquivo de saída
        """
        try:
            # Cria novo workbook
            from openpyxl import Workbook
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Resultados"
            
            # Escreve dados
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), 1):
                for c_idx, value in enumerate(row, 1):
                    new_ws.cell(row=r_idx, column=c_idx, value=value)
                    
            # Aplica formatação especial para linhas de Vendas e Sugestão
            self.apply_excel_formatting(new_ws)
            
            # Salva arquivo
            new_wb.save(output_path)
            
        except Exception as e:
            raise Exception(f"Erro ao exportar resultados: {str(e)}")
            
    def apply_excel_formatting(self, worksheet):
        """
        Aplica formatação do Excel para o novo formato
        Destaca a coluna de Sugestão
        
        Args:
            worksheet: Planilha do openpyxl para aplicar formatação
        """
        # Define cores
        suggestion_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Amarelo claro
        header_color = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cinza claro
        bold_font = Font(bold=True)
        
        # Aplica formatação na coluna de Sugestão (coluna S = 19)
        suggestion_col = 19  # Coluna S
        
        for row_num in range(1, worksheet.max_row + 1):
            # Destaca a coluna de sugestão
            cell = worksheet.cell(row=row_num, column=suggestion_col)
            if cell.value and str(cell.value).strip() != "" and str(cell.value).strip() != "0":
                cell.fill = suggestion_color
                cell.font = bold_font
                
        # Aplica formatação no cabeçalho
        for col_num in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col_num)
            if header_cell.value:
                header_cell.fill = header_color
                header_cell.font = bold_font
                
        # Formata números na coluna de sugestão
        for row_num in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_num, column=suggestion_col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0'
                        
    def get_calculation_summary(self) -> dict:
        """
        Retorna resumo dos cálculos realizados para o novo formato
        
        Returns:
            Dicionário com estatísticas dos cálculos
        """
        summary = {
            'total_rows': len(self.df) if self.df is not None else 0,
            'product_rows': len(self.find_product_rows()) if self.df is not None else 0,
            'suggestions_calculated': 0,
            'columns': len(self.df.columns) if self.df is not None and not self.df.empty else 0,
            'has_suggestion_column': False
        }
        
        if self.df is not None and not self.df.empty:
            # Verifica se há coluna de sugestão
            if len(self.df.columns) > 18:
                summary['has_suggestion_column'] = True
                suggestion_col = self.df.columns[18]
                
                # Conta quantas sugestões foram calculadas
                for idx, row in self.df.iterrows():
                    valor = row[suggestion_col]
                    if pd.notna(valor) and isinstance(valor, (int, float)) and valor > 0:
                        summary['suggestions_calculated'] += 1
        
        return summary
        
    def transpose_to_old_format(self) -> pd.DataFrame:
        """
        Transpõe dados do formato novo para o formato antigo
        """
        print("\n🔄 INICIANDO TRANSPOSIÇÃO DE FORMATO")
        print(f"📊 Dados originais: {len(self.df)} linhas, {len(self.df.columns)} colunas")
        
        if self.df.empty:
            print("❌ DataFrame vazio - não há dados para transpor")
            return pd.DataFrame()
            
        # Log das primeiras linhas para debug
        print("\n📋 PRIMEIRAS 10 LINHAS DO ARQUIVO:")
        for i in range(min(10, len(self.df))):
            if len(self.df.columns) > 0:
                primeira_col = self.df.iloc[i, 0]
                print(f"  Linha {i}: '{primeira_col}'")
        
        # Detecta formato
        shops_found = []
        current_shop = None
        shop_data = {}
        
        print("\n🏪 PROCURANDO LOJAS NO ARQUIVO:")
        
        for idx, row in self.df.iterrows():
            primeira_col = row.iloc[0] if len(row) > 0 else None
            
            if pd.notna(primeira_col):
                primeira_str = str(primeira_col).strip()
                
                # Verifica se é linha de loja
                if 'loja:' in primeira_str.lower() or '(' in primeira_str:
                    shop_code = self.extract_shop_code(primeira_str)
                    if shop_code:
                        current_shop = shop_code
                        shops_found.append(shop_code)
                        shop_data[shop_code] = []
                        print(f"  ✅ Loja encontrada na linha {idx}: '{shop_code}' (linha original: '{primeira_str}')")
                    else:
                        print(f"  ❌ Linha parece ser loja mas não conseguiu extrair código na linha {idx}: '{primeira_str}'")
                        
                # Verifica se é linha de produto
                elif current_shop and self.is_product_line(row):
                    print(f"  📦 Produto encontrado na linha {idx} para loja '{current_shop}'")
                    shop_data[current_shop].append((idx, row))
                    
        print(f"\n📈 RESUMO DA ANÁLISE:")
        print(f"  🏪 Total de lojas encontradas: {len(shops_found)}")
        print(f"  🏪 Lojas: {shops_found}")
        
        for shop, products in shop_data.items():
            print(f"  📦 Loja '{shop}': {len(products)} produtos")
            
        if not shops_found:
            print("❌ ERRO: Nenhuma loja encontrada no arquivo!")
            print("🔍 Verificando possíveis causas:")
            
            # Analisa cada linha procurando padrões
            for idx, row in self.df.head(20).iterrows():
                if len(row) > 0 and pd.notna(row.iloc[0]):
                    linha = str(row.iloc[0]).strip()
                    print(f"    Linha {idx}: '{linha}' - Contém 'loja': {'loja' in linha.lower()}, Contém '(': {'(' in linha}")
                    
            return pd.DataFrame()
            
        # Cria estrutura do formato antigo
        print(f"\n🔧 CRIANDO ESTRUTURA DO FORMATO ANTIGO")
        transposed_data = self.create_old_format_structure(shop_data, shops_found)
        
        print(f"✅ TRANSPOSIÇÃO CONCLUÍDA: {len(transposed_data)} linhas criadas")
        return pd.DataFrame(transposed_data)
        
    def extract_shop_code(self, linha: str) -> Optional[str]:
        """
        Extrai código da loja de uma linha
        """
        patterns = [
            r'\(([^)]+)\)',  # Entre parênteses
            r'loja:\s*([^-\s]+)',  # Após "loja:"
        ]
        
        for pattern in patterns:
            match = re.search(pattern, linha, re.IGNORECASE)
            if match:
                return match.group(1).strip()
                
        return None
        
    def is_product_line(self, row: pd.Series) -> bool:
        """
        Verifica se uma linha contém dados de produto
        """
        # Verifica se tem pelo menos código e descrição
        if len(row) < 4:
            return False
            
        # Coluna C deve ter código (numérico)
        codigo = row.iloc[2] if len(row) > 2 else None
        if pd.isna(codigo):
            return False
            
        try:
            int(str(codigo).strip())
            return True
        except:
            return False
            
    def create_old_format_structure(self, shop_data: Dict, shops_found: List[str]) -> List[Dict]:
        """
        Cria a estrutura do formato antigo
        """
        print("🏗️ CRIANDO ESTRUTURA DO FORMATO ANTIGO")
        
        # Headers do formato antigo
        base_headers = ['Tipo', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 
                       'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada']
        
        # Adiciona colunas das lojas
        all_headers = base_headers + shops_found + ['Qtde Total']
        
        print(f"📋 Headers criados: {all_headers}")
        
        # Coleta todos os produtos únicos
        all_products = {}
        
        for shop, products in shop_data.items():
            print(f"🏪 Processando {len(products)} produtos da loja '{shop}'")
            
            for idx, row in products:
                # Extrai dados do produto
                codigo = str(row.iloc[2]).strip() if len(row) > 2 else ""
                descricao = str(row.iloc[6]).strip() if len(row) > 6 else ""
                
                if codigo and codigo != 'nan':
                    if codigo not in all_products:
                        all_products[codigo] = {
                            'Codigo': codigo,
                            'Descricao': descricao,
                            'Cx c/': str(row.iloc[7]) if len(row) > 7 else "",
                            'Secundario': str(row.iloc[3]) if len(row) > 3 else "",
                            'shops_vendas': {},
                            'shops_entradas': {}
                        }
                    
                    # Pega valores de vendas e entradas
                    vendas = row.iloc[8] if len(row) > 8 else 0  # Coluna I (estoque)
                    entradas = row.iloc[11] if len(row) > 11 else 0  # Coluna L (entrada)
                    
                    all_products[codigo]['shops_vendas'][shop] = vendas if pd.notna(vendas) else 0
                    all_products[codigo]['shops_entradas'][shop] = entradas if pd.notna(entradas) else 0
        
        print(f"📦 Total de produtos únicos encontrados: {len(all_products)}")
        
        # Cria linhas do formato antigo
        result_data = []
        
        for codigo, product_data in all_products.items():
            # Linha de Entradas
            entrada_row = {
                'Tipo': 'Entradas',
                'Codigo': product_data['Codigo'],
                'Descricao': product_data['Descricao'],
                'Cx c/': product_data['Cx c/'],
                'Secundario': product_data['Secundario'],
                'Saldo Local': 0,
                'Invent': 0,
                'Devol.': 0,
                'Dep25': 0,
                'Entrada': 0
            }
            
            # Adiciona valores das lojas para entradas
            total_entradas = 0
            for shop in shops_found:
                valor = product_data['shops_entradas'].get(shop, 0)
                entrada_row[shop] = valor
                total_entradas += valor if isinstance(valor, (int, float)) else 0
                
            entrada_row['Qtde Total'] = total_entradas
            result_data.append(entrada_row)
            
            # Linha de Vendas
            vendas_row = {
                'Tipo': 'Vendas',
                'Codigo': product_data['Codigo'],
                'Descricao': product_data['Descricao'],
                'Cx c/': product_data['Cx c/'],
                'Secundario': product_data['Secundario'],
                'Saldo Local': 0,
                'Invent': 0,
                'Devol.': 0,
                'Dep25': 0,
                'Entrada': 0
            }
            
            # Adiciona valores das lojas para vendas
            total_vendas = 0
            for shop in shops_found:
                valor = product_data['shops_vendas'].get(shop, 0)
                vendas_row[shop] = valor
                total_vendas += valor if isinstance(valor, (int, float)) else 0
                
            vendas_row['Qtde Total'] = total_vendas
            result_data.append(vendas_row)
            
        print(f"✅ Estrutura criada: {len(result_data)} linhas (Entradas + Vendas)")
        return result_data