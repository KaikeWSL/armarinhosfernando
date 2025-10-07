import pandas as pd
import re
from typing import List, Dict, Optional

def sort_stores_custom_order(stores_found):
    """Ordena as lojas na ordem específica solicitada"""
    # Ordem específica das lojas
    ordem_preferida = ['MTZ', 'FL02', 'FL03', 'TAT', 'MOOCA', 'BRAS', 'FL09', 'SMP', 'SBC', 'SAN', 'LP', 'GRU', 'IPI', 'SAM', 'OSA', 'SOR', 'SJC', 'ABDO']
    
    # Separa lojas conhecidas das novas
    lojas_conhecidas = []
    lojas_novas = []
    
    for loja in stores_found:
        if loja in ordem_preferida:
            lojas_conhecidas.append(loja)
        else:
            lojas_novas.append(loja)
    
    # Ordena lojas conhecidas pela ordem preferida
    lojas_ordenadas = []
    for loja_ordem in ordem_preferida:
        if loja_ordem in lojas_conhecidas:
            lojas_ordenadas.append(loja_ordem)
    
    # Adiciona lojas novas no final (após ABDO), ordenadas alfabeticamente
    lojas_novas_ordenadas = sorted(lojas_novas)
    
    return lojas_ordenadas + lojas_novas_ordenadas

class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.df = None
        self.transposed_df = None
        
    def load_excel_file(self):
        """Carrega o arquivo Excel"""
        try:
            self.df = pd.read_excel(self.file_path)
            print(f"📂 Arquivo carregado: {len(self.df)} linhas, {len(self.df.columns)} colunas")
            return True
        except Exception as e:
            print(f"❌ Erro ao carregar arquivo: {e}")
            return False

    def transpose_to_old_format(self):
        """Transpõe dados do novo formato para o formato antigo"""
        if self.df is None:
            print("❌ Nenhum dado carregado")
            return pd.DataFrame()
        
        print("\n🔄 FASE 1: Analisando dados...")
        products_data = []
        stores_found = []
        current_store = None
        
        # Itera pelas linhas do DataFrame
        for idx, row in self.df.iterrows():
            first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            
            # Detecta linha de loja
            if first_col.startswith("Loja:"):
                store_code = self.extract_store_code(first_col)
                if store_code:
                    current_store = store_code
                    if store_code not in stores_found:
                        stores_found.append(store_code)
                    print(f"🏪 Loja: {store_code}")
                continue
            
            # DEBUG específico para produto 1075707
            if "1075707" in str(row.iloc[2]):
                print(f"  🔍 DEBUG linha {idx}: raw_codigo={row.iloc[2]}, req={self.safe_float(row.iloc[12])}, vendas={self.safe_float(row.iloc[14])}")
            
            # Detecta linha de produto
            if current_store and self.is_product_line(row):
                try:
                    product_info = self.extract_product_data(row, current_store)
                    if product_info:
                        products_data.append(product_info)
                        print(f"  📦 {product_info['codigo']} em {current_store}: V={product_info['vendas']}, E={product_info['entradas']}")
                except Exception as e:
                    print(f"⚠️ Erro linha {idx}: {str(e)}")
                    continue

        print(f"\n📊 RESUMO DETECTADO:")
        print(f"🏪 Lojas: {stores_found}")
        print(f"📦 Total de registros produto-loja: {len(products_data)}")
        
        if not stores_found or not products_data:
            print("❌ Nenhum dado válido encontrado")
            return pd.DataFrame()
        
        print("\n🔄 FASE 2: Criando formato antigo...")
        
        # Agrupa produtos por código (SEM DUPLICAÇÃO)
        products_by_code = {}
        for product in products_data:
            codigo = product['codigo']
            if codigo not in products_by_code:
                products_by_code[codigo] = {
                    'secundario': product['secundario'],
                    'familia': product['familia'],
                    'descricao': product['descricao'],
                    'cx_c': product['cx_c'],
                    'invent': product['invent'],
                    'devol': product['devol'],
                    'lojas_vendas': {},
                    'lojas_entradas': {},
                    'lojas_requisicoes': {},
                    'lojas_requisicoes_abdo': {}
                }
            
            # Armazena valores específicos da loja (SOBRESCREVE se já existe)
            loja = product['loja']
            products_by_code[codigo]['lojas_vendas'][loja] = product['vendas']
            products_by_code[codigo]['lojas_entradas'][loja] = product['entradas']
            products_by_code[codigo]['lojas_requisicoes'][loja] = product['requisicoes']
            products_by_code[codigo]['lojas_requisicoes_abdo'][loja] = product['requisicoes_abdo']
        
        print(f"🗂️ Produtos únicos: {len(products_by_code)}")
        
        # Cria DataFrame no formato antigo
        result_data = []
        
        for codigo, info in products_by_code.items():
            # Linha de Entradas
            entrada_row = {
                'Tipo': 'Entradas',
                'Familia': info['familia'],
                'Codigo': codigo,
                'Descricao': info['descricao'],
                'Cx c/': info['cx_c'],
                'Secundario': info['secundario'],
                'Saldo Local': 0,  # Campo padrão
                'Invent': info['invent'],
                'Devol.': info['devol'],
                'Dep25': 0,  # Campo padrão
                'Entrada': sum(info['lojas_entradas'].values()),  # Total de entradas
                'Requisições ABDO': sum(info['lojas_requisicoes_abdo'].values()),  # Total requisições ABDO
                'Requisições': sum(info['lojas_requisicoes'].values())  # Total requisições
            }
            
            # Adiciona valores de entrada para cada loja
            for loja in stores_found:
                entrada_row[loja] = info['lojas_entradas'].get(loja, 0)
            
            result_data.append(entrada_row)
            
            # Linha de Vendas
            vendas_row = {
                'Tipo': 'Vendas',
                'Familia': info['familia'],
                'Codigo': codigo,
                'Descricao': info['descricao'],
                'Cx c/': info['cx_c'],
                'Secundario': info['secundario'],
                'Saldo Local': 0,  # Campo padrão
                'Invent': info['invent'],
                'Devol.': info['devol'],
                'Dep25': 0,  # Campo padrão
                'Entrada': sum(info['lojas_vendas'].values()),  # Total de vendas
                'Requisições ABDO': sum(info['lojas_requisicoes_abdo'].values()),  # Total requisições ABDO
                'Requisições': sum(info['lojas_requisicoes'].values())  # Total requisições
            }
            
            # Adiciona valores de vendas para cada loja
            for loja in stores_found:
                vendas_row[loja] = info['lojas_vendas'].get(loja, 0)
            
            result_data.append(vendas_row)
        
        # Cria DataFrame resultado
        if result_data:
            # Ordena colunas conforme especificado: Tipo, Familia, Codigo, Descricao, Cx c/, Secundario, Saldo Local, Invent, Devol., Dep25, Entrada, Requisições ABDO, Requisições, [lojas]
            base_columns = ['Tipo', 'Familia', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada', 'Requisições ABDO', 'Requisições']
            store_columns = sort_stores_custom_order(stores_found)
            final_columns = base_columns + store_columns
            
            result_df = pd.DataFrame(result_data)
            result_df = result_df.reindex(columns=final_columns, fill_value=0)
            
            # Atualiza o DataFrame da instância
            self.transposed_df = result_df
            
            print(f"📊 Resultado: {len(result_df)} linhas, {len(result_df.columns)} colunas")
            print(f"🏪 Colunas de lojas: {[col for col in result_df.columns if col in stores_found]}")
            
            # Debug: mostra primeiros produtos
            if len(result_df) > 0:
                print(f"\n📋 EXEMPLO - Primeiro produto:")
                primeiro_produto = result_df.iloc[0]
                print(f"  Código: {primeiro_produto['Codigo']}")
                for loja in stores_found:
                    if loja in primeiro_produto:
                        print(f"  {loja}: {primeiro_produto[loja]}")
            
            return result_df
        else:
            return pd.DataFrame()

    def extract_store_code(self, text):
        """Extrai código da loja dos parênteses"""
        match = re.search(r'\(([^)]+)\)', text)
        if match:
            return match.group(1).strip()
        return None

    def is_product_line(self, row):
        """Verifica se a linha contém dados de produto"""
        try:
            # Produto na coluna C (index 2)
            if pd.isna(row.iloc[2]):
                return False
            
            codigo_str = str(row.iloc[2]).strip()
            
            # Verifica se é um código numérico válido
            if not codigo_str or codigo_str == "":
                return False
            
            # Remove pontos decimais se houver
            codigo_clean = codigo_str.replace('.0', '')
            
            return codigo_clean.isdigit() and len(codigo_clean) >= 6
        except:
            return False

    def extract_product_data(self, row, loja):
        """Extrai dados do produto de uma linha"""
        try:
            # Código na coluna C (index 2)
            codigo_raw = row.iloc[2]
            if pd.isna(codigo_raw):
                return None
                
            codigo = str(int(float(codigo_raw))) if isinstance(codigo_raw, (int, float)) else str(codigo_raw)
            
            # Secundário na coluna D (index 3) 
            secundario = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
            if secundario and secundario != "nan":
                secundario = secundario.replace('.0', '')  # Remove .0 se for número
            
            # Entradas (requisições) na coluna R (index 17)
            entradas = self.safe_float(row.iloc[17])
            
            # Familia na coluna F (index 5) - ajustei para a posição correta
            familia = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
            
            # Descrição na coluna G (index 6)
            descricao = str(row.iloc[6]) if pd.notna(row.iloc[6]) else ""
            
            # Cx c/ na coluna H (index 7)
            cx_c = str(row.iloc[7]) if pd.notna(row.iloc[7]) else ""
            if cx_c and cx_c != "nan":
                cx_c = cx_c.replace('.0', '')  # Remove .0 se for número
            
            # Invent na coluna J (index 9)
            invent = self.safe_float(row.iloc[9])
            
            # Devol. na coluna K (index 10)
            devol = self.safe_float(row.iloc[10])
            
            # Vendas na coluna O (index 14)
            vendas = self.safe_float(row.iloc[14])
            
            # Requisições na coluna M (index 12)
            requisicoes = self.safe_float(row.iloc[12])
            
            # Requisições ABDO na coluna N (index 13)
            requisicoes_abdo = self.safe_float(row.iloc[13])
            
            return {
                'codigo': codigo,
                'secundario': secundario,
                'familia': familia,
                'descricao': descricao,
                'cx_c': cx_c,
                'invent': invent,
                'devol': devol,
                'loja': loja,
                'entradas': entradas,
                'vendas': vendas,
                'requisicoes': requisicoes,
                'requisicoes_abdo': requisicoes_abdo
            }
        except Exception as e:
            print(f"❌ Erro ao extrair produto: {e}")
            return None

    def safe_float(self, value):
        """Converte valor para float de forma segura"""
        try:
            if pd.isna(value):
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def get_data_preview(self):
        """Retorna preview dos dados transpostos"""
        if self.transposed_df is not None:
            return self.transposed_df  # Retorna todos os dados
        return pd.DataFrame()

    def calculate_suggestions(self, percentage: float, progress_callback=None):
        """Calcula sugestões com base na porcentagem e insere após cada linha de vendas"""
        if self.transposed_df is None:
            return False
        
        # Identifica colunas de lojas
        info_columns = ['Tipo', 'Familia', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada', 'Requisições ABDO', 'Requisições']
        store_columns = [col for col in self.transposed_df.columns 
                        if col not in info_columns]
        
        # Cria novo DataFrame com sugestões inseridas
        new_data = []
        total_products = len(self.transposed_df[self.transposed_df['Tipo'] == 'Vendas'])
        current_product = 0
        
        # Agrupa por produto (código)
        produtos_unicos = self.transposed_df['Codigo'].unique()
        
        for produto_codigo in produtos_unicos:
            produto_rows = self.transposed_df[self.transposed_df['Codigo'] == produto_codigo]
            
            # Adiciona linha de entradas
            entrada_row = produto_rows[produto_rows['Tipo'] == 'Entradas']
            if not entrada_row.empty:
                new_data.append(entrada_row.iloc[0])
            
            # Adiciona linha de vendas
            vendas_row = produto_rows[produto_rows['Tipo'] == 'Vendas']
            if not vendas_row.empty:
                vendas_data = vendas_row.iloc[0]
                new_data.append(vendas_data)
                
                # Cria e adiciona linha de sugestão (DEPOIS das vendas)
                suggestion_row = vendas_data.copy()
                suggestion_row['Tipo'] = 'Sugestão'
                
                # Aplica porcentagem nas colunas de lojas
                for col in store_columns:
                    original_value = vendas_data[col]
                    if pd.notna(original_value) and original_value > 0:
                        suggestion_row[col] = int(original_value * (1 + percentage / 100))
                
                # Recalcula total de entrada (soma das lojas)
                suggestion_row['Entrada'] = sum([suggestion_row[col] for col in store_columns if pd.notna(suggestion_row[col])])
                
                new_data.append(suggestion_row)
                
                current_product += 1
                # Callback de progresso
                if progress_callback:
                    progress_callback(int(current_product / total_products * 100))
        
        # Atualiza o DataFrame
        if new_data:
            self.transposed_df = pd.DataFrame(new_data).reset_index(drop=True)
        
        return True

    def export_results(self, output_path: str):
        """Exporta resultados para Excel"""
        if self.transposed_df is None:
            return False
        
        try:
            self.transposed_df.to_excel(output_path, index=False)
            return True
        except Exception as e:
            print(f"❌ Erro ao exportar: {e}")
            return False

    def get_calculation_summary(self):
        """Retorna resumo"""
        if self.transposed_df is None:
            return {'message': 'Nenhum dado processado'}
        
        vendas_count = len(self.transposed_df[self.transposed_df['Tipo'] == 'Vendas'])
        entradas_count = len(self.transposed_df[self.transposed_df['Tipo'] == 'Entradas'])
        sugestoes_count = len(self.transposed_df[self.transposed_df['Tipo'] == 'Sugestão'])
        
        return {
            'total_rows': len(self.transposed_df),
            'vendas_rows': vendas_count,
            'entradas_rows': entradas_count,
            'suggestion_rows': sugestoes_count,
            'produtos_unicos': vendas_count
        }
    
    def process_file(self):
        """Processa o arquivo completo"""
        try:
            self.load_excel_file()
            self.transpose_to_old_format()
            return True
        except Exception as e:
            print(f"Erro no processamento: {e}")
            return False
    
    def get_preview_data(self) -> pd.DataFrame:
        """Retorna dados para preview"""
        if self.transposed_df is not None:
            return self.transposed_df  # Retorna todos os dados
        return pd.DataFrame()
    
    def apply_percentage_adjustment(self, percentage: float):
        """Aplica ajuste percentual nas vendas e reorganiza com sugestões no meio"""
        if self.transposed_df is None:
            return
        
        # Identifica colunas de lojas
        info_columns = ['Tipo', 'Familia', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada', 'Requisições ABDO', 'Requisições']
        store_columns = [col for col in self.transposed_df.columns 
                        if col not in info_columns]
        
        # Cria novo DataFrame com sugestões inseridas no meio
        new_data = []
        
        # Agrupa por produto (código)
        produtos_unicos = self.transposed_df['Codigo'].unique()
        
        for produto_codigo in produtos_unicos:
            produto_rows = self.transposed_df[self.transposed_df['Codigo'] == produto_codigo]
            
            # Adiciona linha de entradas
            entrada_row = produto_rows[produto_rows['Tipo'] == 'Entradas']
            if not entrada_row.empty:
                new_data.append(entrada_row.iloc[0])
            
            # Adiciona linha de vendas
            vendas_row = produto_rows[produto_rows['Tipo'] == 'Vendas']
            if not vendas_row.empty:
                vendas_data = vendas_row.iloc[0]
                new_data.append(vendas_data)
                
                # Cria e adiciona linha de sugestão (DEPOIS das vendas)
                suggestion_row = vendas_data.copy()
                suggestion_row['Tipo'] = 'Sugestão'
                
                # Aplica porcentagem nas colunas de lojas
                for col in store_columns:
                    if pd.notna(vendas_data[col]) and vendas_data[col] != 0:
                        suggestion_row[col] = int(vendas_data[col] * (1 + percentage / 100))
                
                # Recalcula total de entrada (soma das lojas)
                suggestion_row['Entrada'] = sum([suggestion_row[col] for col in store_columns if pd.notna(suggestion_row[col])])
                
                new_data.append(suggestion_row)
        
        # Atualiza o DataFrame
        if new_data:
            self.transposed_df = pd.DataFrame(new_data).reset_index(drop=True)
    
    def export_to_excel(self, output_path: str, familia_filter=None):
        """Exporta o resultado para Excel no formato antigo com cabeçalho formatado e totais"""
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime
            
            # Obtém dados com totais inclusos
            df_to_export = self.get_data_with_totals()
            
            # Filtra por família se especificado
            if familia_filter and familia_filter != "Todas":
                # Mantém as linhas de totais
                mask = (df_to_export['Familia'] == familia_filter) | (df_to_export['Tipo'].str.contains('Quantidade Total', na=False))
                df_to_export = df_to_export[mask]
            
            # Cria novo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Formato Antigo"
            
            # Linha 1: Título principal
            ws['A1'] = "Relatório Entrada X Venda de Saldo Estoque Lojas"
            ws['A1'].font = Font(bold=True, size=14)
            
            # Linha 2: Nome da empresa
            ws['A2'] = "Armarinhos Fernando Ltda Matriz"
            ws['A2'].font = Font(bold=True, size=12)
            
            # Linha 3: Vazia
            
            # Linha 4: Família
            familia_text = familia_filter if familia_filter and familia_filter != "Todas" else "TODAS"
            ws['A4'] = "Familia"
            ws['B4'] = familia_text
            ws['A4'].font = Font(bold=True)
            
            # Linha 5: Período
            data_atual = datetime.now().strftime("%d/%m/%Y")
            ws['A5'] = "Saida Entre"
            ws['B5'] = f"22/03/2024 a {data_atual}"
            ws['A5'].font = Font(bold=True)
            
            # Linha 6: Vazia
            
            # Linha 7: Cabeçalhos dos dados (inicia na linha 7)
            start_row = 7
            
            # Escreve cabeçalhos
            for c_idx, col_name in enumerate(df_to_export.columns, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Escreve dados a partir da linha 8
            for r_idx, row in enumerate(dataframe_to_rows(df_to_export, index=False, header=False), start_row + 1):
                # Identifica o tipo da linha (primeira coluna)
                tipo_linha = row[0] if row and len(row) > 0 else ""
                
                # Define cor da linha baseada no tipo
                fill_color = None
                if isinstance(tipo_linha, str):
                    if 'Vendas' in tipo_linha or 'Vendida' in tipo_linha:
                        fill_color = PatternFill(start_color="faed7a", end_color="faed7a", fill_type="solid")  # Amarelo
                    elif 'Sugestão' in tipo_linha or 'Solicitada' in tipo_linha:
                        fill_color = PatternFill(start_color="aefdb1", end_color="aefdb1", fill_type="solid")  # Verde claro
                    elif 'Entradas' in tipo_linha or 'Entrada' in tipo_linha:
                        fill_color = PatternFill(start_color="e5f3ff", end_color="e5f3ff", fill_type="solid")  # Azul claro
                    elif 'Quantidade Total' in tipo_linha:
                        if 'Vendida' in tipo_linha:
                            fill_color = PatternFill(start_color="d85a5a", end_color="d85a5a", fill_type="solid")  # Laranja
                        elif 'Solicitada' in tipo_linha:
                            fill_color = PatternFill(start_color="d85a5a", end_color="d85a5a", fill_type="solid")  # Verde mais escuro
                        elif 'Entrada' in tipo_linha:
                            fill_color = PatternFill(start_color="d85a5a", end_color="d85a5a", fill_type="solid")  # Azul mais escuro
                
                # Aplica cor em todas as células da linha
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if fill_color:
                        cell.fill = fill_color
            
            # Ajusta largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salva arquivo
            wb.save(output_path)
            print(f"Arquivo exportado com totais: {output_path}")
            
        except Exception as e:
            print(f"Erro ao exportar: {e}")
            raise
    
    def get_data_with_totals(self):
        """Retorna os dados transpostos com linhas de totais inclusos"""
        if self.transposed_df is None:
            return pd.DataFrame()
            
        df = self.transposed_df.copy()
        
        # Identifica colunas das lojas (separadas das administrativas)
        lojas_cols = []
        admin_cols = ['Cx c/', 'Secundario', 'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entradas', 'Requisições ABDO', 'Requisições']
        
        for col in df.columns:
            if col not in ['Tipo', 'Familia', 'Codigo', 'Descricao'] and df[col].dtype in ['int64', 'float64']:
                if col not in admin_cols:
                    lojas_cols.append(col)  # Colunas das lojas
        
        # Todas as colunas numéricas (para o Total da linha)
        all_numeric_cols = lojas_cols + [c for c in admin_cols if c in df.columns]
        
        # Adiciona coluna Total para cada linha (soma TODAS as colunas numéricas)
        df['Total'] = 0.0
        for i in range(len(df)):
            total_linha = 0.0
            for col in all_numeric_cols:
                valor = df.iloc[i][col]
                if pd.notna(valor) and isinstance(valor, (int, float)):
                    try:
                        total_linha += float(valor)
                    except (ValueError, TypeError, OverflowError):
                        continue
            # Converte para int se for número inteiro
            total_linha = round(total_linha, 2)
            if total_linha.is_integer():
                df.loc[i, 'Total'] = int(total_linha)
            else:
                df.loc[i, 'Total'] = total_linha
        
        # Calcula totais por tipo
        totals_data = {}
        for tipo in ['VENDAS', 'SUGESTÃO', 'ENTRADAS']:
            mask = df['Tipo'].str.upper().str.contains(tipo, na=False)
            if mask.any():
                totals_data[tipo] = df[mask]
        
        # Adiciona linha em branco
        empty_row = {col: "" for col in df.columns}
        df = pd.concat([df, pd.DataFrame([empty_row])], ignore_index=True)
        
        # Adiciona linhas de totais
        for tipo, nome_total in [('VENDAS', 'Quantidade Total Vendida'), 
                                ('SUGESTÃO', 'Quantidade Total Solicitada'),
                                ('ENTRADAS', 'Quantidade Total Entrada')]:
            if tipo in totals_data and not totals_data[tipo].empty:
                total_row = {col: "" for col in df.columns}
                total_row['Tipo'] = nome_total
                
                # Soma APENAS colunas das lojas + Total (não inclui colunas administrativas)
                for col in lojas_cols + ['Total']:
                    if col in totals_data[tipo].columns and col in lojas_cols:
                        # Para colunas das lojas
                        try:
                            soma = totals_data[tipo][col].sum()
                            soma_float = float(soma)
                            total_row[col] = int(soma_float) if soma_float.is_integer() and soma_float != 0 else (soma_float if soma_float != 0 else 0)
                        except (ValueError, TypeError, OverflowError):
                            total_row[col] = 0
                    elif col == 'Total':
                        # Para a coluna Total, soma APENAS as colunas das lojas
                        total_soma = 0.0
                        for c in lojas_cols:
                            if c in totals_data[tipo].columns:
                                try:
                                    valor_col = totals_data[tipo][c].sum()
                                    total_soma += float(valor_col)
                                except (ValueError, TypeError, OverflowError):
                                    continue
                        total_row[col] = int(total_soma) if total_soma.is_integer() else round(total_soma, 2)
                
                df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        
        return df
            
    def get_available_families(self):
        """Retorna lista de famílias disponíveis"""
        if self.transposed_df is None:
            return []
        
        families = self.transposed_df['Familia'].unique()
        # Remove valores vazios e ordena
        families = [f for f in families if f and str(f).strip() != "" and str(f) != "nan"]
        return sorted(families)