"""
Processador de arquivos Excel - Versão Transposição
Responsável por carregar, processar e transpor planilhas Excel
Converte formato novo (lojas em grupos) para formato antigo (colunas Entradas/Vendas)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
from typing import Optional, Tuple, List
from calculator import CalculationEngine

class ExcelProcessor:
    def __init__(self, file_path: str):
        """
        Inicializa o processador Excel
        
        Args:
            file_path: Caminho para o arquivo Excel
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.df = None
        self.original_df = None
        self.calculation_engine = CalculationEngine()
        self.transposed_df = None
        self.load_excel_file()
        
    def load_excel_file(self):
        """Carrega o arquivo Excel"""
        try:
            # Carrega o workbook com openpyxl
            self.workbook = load_workbook(self.file_path)
            
            # Pega a primeira planilha
            sheet_name = self.workbook.sheetnames[0]
            self.worksheet = self.workbook[sheet_name]
            
            # Carrega dados com pandas - sem header para preservar estrutura
            self.df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            self.original_df = self.df.copy()
            
            print(f"✅ Arquivo carregado: {len(self.df)} linhas, {len(self.df.columns)} colunas")
            
            # Tenta fazer a transposição imediatamente
            self.transposed_df = self.transpose_to_old_format()
            
        except Exception as e:
            raise Exception(f"Erro ao carregar arquivo Excel: {str(e)}")
            
    def transpose_to_old_format(self):
        """
        Transpõe dados do novo formato para o formato antigo
        CORRIGIDO: Não soma valores, mantém separado por produto
        """
        print("\n🔄 === TRANSPOSIÇÃO PARA FORMATO ANTIGO (CORRIGIDA) ===")
        
        if self.df is None or self.df.empty:
            print("❌ DataFrame vazio")
            return pd.DataFrame()
        
        try:
            # Estruturas para armazenar dados
            stores = []
            store_data = {}
            current_store = None
            
            print(f"📊 Analisando {len(self.df)} linhas...")
            
            # Primeiro passo: identificar lojas e produtos
            for idx, row in self.df.iterrows():
                if len(row) == 0:
                    continue
                    
                first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                
                # Detecta cabeçalho de loja
                if first_col.startswith("Loja:"):
                    store_code = self.extract_store_from_parentheses(first_col)
                    if store_code:
                        current_store = store_code
                        if current_store not in stores:
                            stores.append(current_store)
                            store_data[current_store] = []
                        print(f"🏪 Loja detectada: '{current_store}' da linha: '{first_col}'")
                    else:
                        print(f"⚠️ Não conseguiu extrair loja de: '{first_col}'")
                    continue
                
                # Pula linhas vazias ou totais
                if not first_col or "total" in first_col.lower() or len(first_col) < 3:
                    continue
                
                # Se temos loja atual e esta linha parece ser produto
                if current_store and len(row) >= 9:
                    try:
                        # Mapeia colunas do novo formato
                        zona = first_col  # Coluna A (ex: "03-Bairro")
                        codigo = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                        secundario = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
                        descricao = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
                        requisicoes = self.safe_float(row.iloc[4])
                        vendas = self.safe_float(row.iloc[5])
                        estoque = self.safe_float(row.iloc[6])
                        pedidos = self.safe_float(row.iloc[7])
                        entradas = self.safe_float(row.iloc[8])
                        
                        # Só processa se tem código válido
                        if codigo and len(codigo) > 3:
                            product = {
                                'zona': zona,
                                'codigo': codigo,
                                'secundario': secundario,
                                'descricao': descricao,
                                'requisicoes': requisicoes,
                                'vendas': vendas,
                                'estoque': estoque,
                                'pedidos': pedidos,
                                'entradas': entradas
                            }
                            store_data[current_store].append(product)
                            print(f"  📦 {current_store}: {codigo} - V={vendas}, E={entradas}")
                            
                    except Exception as e:
                        print(f"⚠️ Erro na linha {idx}: {str(e)}")
                        continue
            
            print(f"\n📊 Resumo:")
            print(f"🏪 Lojas encontradas: {stores}")
            for store, products in store_data.items():
                print(f"  {store}: {len(products)} produtos")
            
            if not stores or not any(store_data.values()):
                print("❌ Nenhum dado válido encontrado")
                return pd.DataFrame()
            
            # Coleta todos os produtos únicos SEM SOMAR valores
            all_products = {}
            
            for store, products in store_data.items():
                for product in products:
                    codigo = product['codigo']
                    # Chave única: codigo + loja para não misturar dados
                    key = f"{codigo}_{store}"
                    
                    if codigo not in all_products:
                        all_products[codigo] = {
                            'descricao': product['descricao'],
                            'secundario': product['secundario'],
                            'vendas_por_loja': {},
                            'entradas_por_loja': {}
                        }
                    
                    # CORRIGIDO: Armazena valor específico da loja, não soma
                    all_products[codigo]['vendas_por_loja'][store] = product['vendas']
                    all_products[codigo]['entradas_por_loja'][store] = product['entradas']
            
            print(f"🎯 Produtos únicos: {len(all_products)}")
            
            # Debug: mostra dados por produto
            for codigo, data in list(all_products.items())[:3]:  # Mostra primeiros 3
                print(f"📦 Produto {codigo}:")
                print(f"  Vendas: {data['vendas_por_loja']}")
                print(f"  Entradas: {data['entradas_por_loja']}")
            
            # Cria DataFrame no formato antigo
            result_rows = []
            
            for codigo, data in all_products.items():
                # Linha de ENTRADAS
                entrada_row = {
                    'Tipo': 'Entradas',
                    'Codigo': codigo,
                    'Descricao': data['descricao'],
                    'Cx c/': '',
                    'Secundario': data['secundario'],
                    'Saldo Local': 0,
                    'Invent': 0,
                    'Devol.': 0,
                    'Dep25': 0,
                    'Entrada': 0
                }
                
                # Linha de VENDAS
                venda_row = {
                    'Tipo': 'Vendas',
                    'Codigo': codigo,
                    'Descricao': data['descricao'],
                    'Cx c/': '',
                    'Secundario': data['secundario'],
                    'Saldo Local': 0,
                    'Invent': 0,
                    'Devol.': 0,
                    'Dep25': 0,
                    'Entrada': 0
                }
                
                # Adiciona dados de cada loja (SEM SOMAR)
                total_entradas = 0
                total_vendas = 0
                
                for store in stores:
                    # Pega valor específico da loja para este produto
                    entradas_val = data['entradas_por_loja'].get(store, 0)
                    vendas_val = data['vendas_por_loja'].get(store, 0)
                    
                    entrada_row[store] = entradas_val
                    venda_row[store] = vendas_val
                    
                    total_entradas += entradas_val
                    total_vendas += vendas_val
                
                # Coluna de total
                entrada_row['Qtde Total'] = total_entradas
                venda_row['Qtde Total'] = total_vendas
                
                result_rows.append(entrada_row)
                result_rows.append(venda_row)
            
            result_df = pd.DataFrame(result_rows)
            print(f"✅ Transposição concluída: {len(result_df)} linhas, {len(result_df.columns)} colunas")
            print(f"🏪 Colunas de lojas criadas: {[col for col in result_df.columns if col in stores]}")
            
            return result_df
            
        except Exception as e:
            print(f"❌ Erro na transposição: {str(e)}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def extract_store_from_parentheses(self, text):
        """
        Extrai código da loja dos parênteses
        Ex: "Loja: 06-ACD (CD) - 1 item" -> "CD"
        Ex: "Loja: 09-L1101 (MTZ) - 29 items" -> "MTZ"
        Ex: "Loja: 10-L1201 (FL02) - 5 items" -> "FL02"
        """
        if pd.isna(text):
            return None
            
        text_str = str(text).strip()
        print(f"🔍 Extraindo loja de: '{text_str}'")
        
        # Procura por texto entre parênteses
        match = re.search(r'\(([^)]+)\)', text_str)
        if match:
            store_code = match.group(1).strip()
            print(f"✅ Código extraído: '{store_code}'")
            return store_code
        else:
            print(f"❌ Não encontrou parênteses em: '{text_str}'")
            return None
    
    def safe_float(self, value):
        """Converte valor para float de forma segura"""
        if pd.isna(value):
            return 0.0
        try:
            clean_value = str(value).replace(',', '').replace(' ', '')
            return float(clean_value) if clean_value else 0.0
        except:
            return 0.0
    
    def get_data_preview(self) -> pd.DataFrame:
        """Retorna preview dos dados transpostos"""
        if self.transposed_df is not None and not self.transposed_df.empty:
            print(f"📊 Preview: {len(self.transposed_df)} linhas (formato antigo), {len(self.transposed_df.columns)} colunas")
            return self.transposed_df.copy()
        else:
            print("⚠️ Nenhum dado transposto disponível")
            return pd.DataFrame({"Aviso": ["Não foi possível transpor os dados"]})
    
    def calculate_suggestions(self, percentage: float, progress_callback=None) -> bool:
        """Aplica porcentagem às linhas de Vendas"""
        try:
            if self.transposed_df is None or self.transposed_df.empty:
                raise Exception("Nenhum dado transposto disponível")
            
            print(f"🧮 Aplicando {percentage}% às vendas...")
            
            # Identifica colunas de lojas (excluindo colunas fixas)
            fixed_cols = ['Tipo', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 
                         'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada', 'Qtde Total']
            store_cols = [col for col in self.transposed_df.columns if col not in fixed_cols]
            
            suggestions_added = 0
            
            # Para cada linha de Vendas, cria linha de Sugestão
            new_rows = []
            for idx, row in self.transposed_df.iterrows():
                new_rows.append(row.copy())
                
                if row['Tipo'] == 'Vendas':
                    # Cria linha de sugestão
                    suggestion_row = row.copy()
                    suggestion_row['Tipo'] = 'Sugestão'
                    
                    # Aplica porcentagem nas colunas de lojas
                    for col in store_cols:
                        original_val = row[col]
                        if isinstance(original_val, (int, float)) and original_val > 0:
                            suggestion_row[col] = original_val * (1 + percentage / 100)
                    
                    # Recalcula total
                    total = sum(suggestion_row[col] for col in store_cols if isinstance(suggestion_row[col], (int, float)))
                    suggestion_row['Qtde Total'] = total
                    
                    new_rows.append(suggestion_row)
                    suggestions_added += 1
            
            # Atualiza DataFrame
            self.transposed_df = pd.DataFrame(new_rows).reset_index(drop=True)
            
            print(f"✅ {suggestions_added} sugestões adicionadas")
            return True
            
        except Exception as e:
            print(f"❌ Erro no cálculo: {str(e)}")
            return False
    
    def export_results(self, output_path: str):
        """Exporta resultados transpostos"""
        try:
            if self.transposed_df is None or self.transposed_df.empty:
                raise Exception("Nenhum dado para exportar")
            
            self.transposed_df.to_excel(output_path, index=False)
            print(f"✅ Dados exportados para: {output_path}")
            
        except Exception as e:
            raise Exception(f"Erro ao exportar: {str(e)}")
    
    def get_calculation_summary(self) -> dict:
        """Retorna resumo dos dados"""
        if self.transposed_df is None:
            return {'message': 'Nenhum dado processado'}
        
        vendas_count = len(self.transposed_df[self.transposed_df['Tipo'] == 'Vendas'])
        entradas_count = len(self.transposed_df[self.transposed_df['Tipo'] == 'Entradas'])
        sugestoes_count = len(self.transposed_df[self.transposed_df['Tipo'] == 'Sugestão'])
        
        return {
            'total_rows': len(self.transposed_df),
            'vendas_rows': vendas_count,
            'entradas_rows': entradas_count,
            'suggestion_rows': sugestoes_count,
            'produtos_unicos': vendas_count,  # Cada produto tem 1 linha de venda
        }