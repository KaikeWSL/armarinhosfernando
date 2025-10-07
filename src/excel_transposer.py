"""
Transpositor de Excel - Converte novo formato para formato antigo
Converte dados de lojas em linhas verticais para colunas horizontais
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from typing import Dict, List, Tuple, Optional

def sort_stores_custom_order(stores_found):
    """Ordena as lojas na ordem específica solicitada"""
    # Ordem específica das lojas
    ordem_preferida = ['MTZ', 'FL02', 'FL03', 'TAT', 'MOOCA', 'BRAS', 'FL09', 'SMP', 'SBC', 'SAN', 'LP', 'GRU', 'IPI', 'SAM', 'OSA', 'SOR', 'SJC', 'ABDO']
    
    # Separa lojas conhecidas das novas
    lojas_conhecidas = []
    lojas_novas = []
    
    for loja in stores_found:
        if loja in ordem_preferida:
            lojas_conhecidas.append(loja)
        else:
            lojas_novas.append(loja)
    
    # Ordena lojas conhecidas pela ordem preferida
    lojas_ordenadas = []
    for loja_ordem in ordem_preferida:
        if loja_ordem in lojas_conhecidas:
            lojas_ordenadas.append(loja_ordem)
    
    # Adiciona lojas novas no final (após ABDO), ordenadas alfabeticamente
    lojas_novas_ordenadas = sorted(lojas_novas)
    
    return lojas_ordenadas + lojas_novas_ordenadas

class ExcelTransposer:
    def __init__(self, file_path: str):
        """
        Inicializa o transpositor Excel
        
        Args:
            file_path: Caminho para o arquivo Excel no novo formato
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.raw_data = None
        self.transposed_data = None
        self.stores_mapping = {}
        self.products_data = {}
        
    def load_file(self):
        """Carrega o arquivo Excel no novo formato"""
        try:
            self.workbook = load_workbook(self.file_path)
            
            # Usa a primeira planilha
            sheet_name = self.workbook.sheetnames[0]
            self.worksheet = self.workbook[sheet_name]
            
            # Carrega dados brutos
            self.raw_data = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            
            print(f"Arquivo carregado: {len(self.raw_data)} linhas, {len(self.raw_data.columns)} colunas")
            
        except Exception as e:
            raise Exception(f"Erro ao carregar arquivo: {str(e)}")
    
    def extract_store_info(self, store_line: str) -> Tuple[str, str]:
        """
        Extrai informações da loja a partir da linha de cabeçalho
        
        Args:
            store_line: Linha como "Loja: 06-ACD (CD) - 1 item"
            
        Returns:
            Tupla com (código_loja, nome_completo)
        """
        # Padrão: "Loja: XX-YYYY (ZZZZ)"
        pattern = r'Loja:\s*(\d+-\w+)\s*\(([^)]+)\)'
        match = re.search(pattern, str(store_line))
        
        if match:
            codigo = match.group(1)  # Ex: "06-ACD"
            sigla = match.group(2)   # Ex: "CD"
            return codigo, sigla
        
        return None, None
    
    def parse_new_format(self):
        """Analisa o novo formato e extrai dados dos produtos por loja"""
        current_store = None
        current_store_code = None
        
        for idx, row in self.raw_data.iterrows():
            # Verifica se é uma linha de loja
            cell_a = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
            
            if "Loja:" in cell_a:
                # Nova loja encontrada
                store_code, store_name = self.extract_store_info(cell_a)
                if store_code:
                    current_store = store_name
                    current_store_code = store_code
                    print(f"Loja encontrada: {store_code} ({store_name})")
                continue
            
            # Verifica se é uma linha de produto (tem código na coluna C)
            if current_store and pd.notna(row.iloc[2]):  # Coluna C = Código
                self.process_product_row(row, current_store, current_store_code)
    
    def process_product_row(self, row: pd.Series, store_name: str, store_code: str):
        """Processa uma linha de produto"""
        try:
            # Extrai dados do produto (novo formato)
            zona = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
            codigo = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""  # Coluna C
            secundario = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""  # Coluna D
            if secundario and secundario != "nan":
                secundario = secundario.replace('.0', '')  # Remove .0 se for número
            familia = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""  # Coluna E
            descricao = str(row.iloc[6]) if pd.notna(row.iloc[6]) else ""  # Coluna G
            cx_c = str(row.iloc[7]) if pd.notna(row.iloc[7]) else ""  # Coluna H
            estoque = float(row.iloc[8]) if pd.notna(row.iloc[8]) and str(row.iloc[8]).replace('.','').replace('-','').isdigit() else 0  # Coluna I
            inventario = float(row.iloc[9]) if pd.notna(row.iloc[9]) and str(row.iloc[9]).replace('.','').replace('-','').isdigit() else 0  # Coluna J
            devolucao = float(row.iloc[10]) if pd.notna(row.iloc[10]) and str(row.iloc[10]).replace('.','').replace('-','').isdigit() else 0  # Coluna K
            entrada = float(row.iloc[11]) if pd.notna(row.iloc[11]) and str(row.iloc[11]).replace('.','').replace('-','').isdigit() else 0  # Coluna L
            
            # Identifica se é linha de vendas ou entradas baseado nos valores
            if entrada > 0:
                tipo = "Entradas"
                valor = entrada
            elif estoque > 0:
                tipo = "Vendas"
                valor = estoque
            else:
                return  # Pula linhas sem dados relevantes
            
            # Cria chave única do produto
            product_key = f"{codigo}_{secundario}"
            
            # Inicializa produto se não existir
            if product_key not in self.products_data:
                self.products_data[product_key] = {
                    'codigo': codigo,
                    'descricao': descricao,
                    'cx_c': cx_c,
                    'secundario': secundario,
                    'familia': familia,
                    'vendas_por_loja': {},
                    'entradas_por_loja': {}
                }
            
            # Adiciona dados da loja
            if tipo == "Vendas":
                self.products_data[product_key]['vendas_por_loja'][store_code] = valor
            else:
                self.products_data[product_key]['entradas_por_loja'][store_code] = valor
                
            print(f"Produto {codigo} - {tipo}: {valor} na loja {store_code}")
            
        except Exception as e:
            print(f"Erro ao processar linha de produto: {e}")
    
    def create_old_format(self):
        """Cria o DataFrame no formato antigo"""
        # Define colunas do formato antigo
        base_columns = [
            'Tipo', 'Familia', 'Codigo', 'Descricao', 'Cx c/', 'Secundario', 
            'Saldo Local', 'Invent', 'Devol.', 'Dep25', 'Entrada', 'Requisições ABDO', 'Requisições'
        ]
        
        # Adiciona colunas das lojas (ordenadas)
        all_stores = set()
        for product in self.products_data.values():
            all_stores.update(product['vendas_por_loja'].keys())
            all_stores.update(product['entradas_por_loja'].keys())
        
        store_columns = sort_stores_custom_order(list(all_stores))
        columns = base_columns + store_columns + ['Qtde Total']
        
        # Cria lista de linhas
        rows = []
        
        for product_key, product in self.products_data.items():
            # Linha de ENTRADAS
            entrada_row = [
                'Entradas',
                product['familia'],
                product['codigo'],
                product['descricao'],
                product['cx_c'],
                product['secundario'],
                0,  # Saldo Local
                0,  # Inventário
                0,  # Devolução
                0,  # Dep25
                0,  # Entrada
                0,  # Requisições ABDO
                0   # Requisições
            ]
            
            # Adiciona valores por loja para entradas
            total_entradas = 0
            for store in store_columns:
                valor = product['entradas_por_loja'].get(store, 0)
                entrada_row.append(valor)
                total_entradas += valor
            
            entrada_row.append(total_entradas)  # Qtde Total
            rows.append(entrada_row)
            
            # Linha de VENDAS
            venda_row = [
                'Vendas',
                product['familia'],
                product['codigo'],
                product['descricao'],
                product['cx_c'],
                product['secundario'],
                0,  # Saldo Local
                0,  # Inventário
                0,  # Devolução
                0,  # Dep25
                0,  # Entrada
                0,  # Requisições ABDO
                0   # Requisições
            ]
            
            # Adiciona valores por loja para vendas
            total_vendas = 0
            for store in store_columns:
                valor = product['vendas_por_loja'].get(store, 0)
                venda_row.append(valor)
                total_vendas += valor
            
            venda_row.append(total_vendas)  # Qtde Total
            rows.append(venda_row)
        
        # Cria DataFrame
        self.transposed_data = pd.DataFrame(rows, columns=columns)
        print(f"Formato antigo criado: {len(self.transposed_data)} linhas, {len(self.transposed_data.columns)} colunas")
    
    def apply_percentage_adjustment(self, percentage: float):
        """Aplica ajuste percentual nas vendas"""
        if self.transposed_data is None:
            return
        
        # Identifica colunas de lojas (após 'Entrada')
        entrada_col_idx = self.transposed_data.columns.get_loc('Entrada')
        store_columns = self.transposed_data.columns[entrada_col_idx + 1:-1]  # Exclui 'Qtde Total'
        
        # Processa cada linha de vendas
        for idx, row in self.transposed_data.iterrows():
            if row['Tipo'] == 'Vendas':
                # Insere linha de sugestão
                suggestion_row = row.copy()
                suggestion_row['Tipo'] = 'Sugestão'
                
                # Aplica porcentagem nas colunas de lojas
                total_sugestao = 0
                for col in store_columns:
                    original_value = row[col]
                    if pd.notna(original_value) and original_value > 0:
                        adjusted_value = original_value * (1 + percentage / 100)
                        suggestion_row[col] = adjusted_value
                        total_sugestao += adjusted_value
                
                suggestion_row['Qtde Total'] = total_sugestao
                
                # Insere a linha de sugestão após a linha de vendas
                self.transposed_data = pd.concat([
                    self.transposed_data.iloc[:idx + 1],
                    pd.DataFrame([suggestion_row]),
                    self.transposed_data.iloc[idx + 1:]
                ]).reset_index(drop=True)
        
        print(f"Ajuste de {percentage}% aplicado")
    
    def export_to_excel(self, output_path: str):
        """Exporta o resultado para Excel no formato antigo"""
        try:
            # Cria novo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Formato Antigo"
            
            # Escreve dados
            for r_idx, row in enumerate(dataframe_to_rows(self.transposed_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Aplica formatação
            self.apply_excel_formatting(ws)
            
            # Salva arquivo
            wb.save(output_path)
            print(f"Arquivo exportado: {output_path}")
            
        except Exception as e:
            raise Exception(f"Erro ao exportar: {str(e)}")
    
    def apply_excel_formatting(self, worksheet):
        """Aplica formatação ao Excel"""
        # Define cores
        entrada_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        venda_fill = PatternFill(start_color="FFF2E8", end_color="FFF2E8", fill_type="solid")
        sugestao_fill = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Formata cabeçalho
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Formata linhas de dados
        for row in worksheet.iter_rows(min_row=2):
            tipo_cell = row[0]
            if tipo_cell.value == "Entradas":
                for cell in row:
                    cell.fill = entrada_fill
            elif tipo_cell.value == "Vendas":
                for cell in row:
                    cell.fill = venda_fill
            elif tipo_cell.value == "Sugestão":
                for cell in row:
                    cell.fill = sugestao_fill
                    cell.font = Font(bold=True)
    
    def get_preview_data(self) -> pd.DataFrame:
        """Retorna dados para preview"""
        if self.transposed_data is not None:
            return self.transposed_data.head(20)  # Primeiras 20 linhas
        return pd.DataFrame()
    
    def process_file(self):
        """Processa o arquivo completo"""
        self.load_file()
        self.parse_new_format()
        self.create_old_format()
        return True